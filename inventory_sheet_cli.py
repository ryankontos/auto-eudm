#!/usr/bin/env python3
"""Guided importer for Sydney inventory tracking workbooks.

The importer reads a dated sheet, previews the resulting device changes, then
either stops (--dry-run), submits real DWP requests, or submits fully local
simulation requests (--simulate). It never submits before showing the preview
and receiving a final confirmation.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass, replace
from datetime import date, datetime
import getpass
import os
from pathlib import Path
import re
import sys
import urllib.parse
from typing import Any, Iterable

import automate_device_request as dwp


NEW_STOCK = "Deployed - New Stock"
EXISTING_STOCK = "Deployed - Existing Stock"
PENDING_RETURN = "Pending Return"
FILE_PREFIX = "Inventory Tracking - Sydney"


@dataclass(frozen=True)
class SheetRow:
    row_number: int
    deployment_date: date
    username: str | None
    username_from_email: bool
    new_serial: str | None
    old_serial: str | None
    marked_red: bool


@dataclass(frozen=True)
class Action:
    group: str
    row_number: int
    username: str
    serial: str
    status: str


@dataclass(frozen=True)
class Outcome:
    action: Action
    request_id: str | None
    order_id: str | None
    error: str | None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("#"):
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    return text


def looks_like_serial(value: str | None) -> bool:
    """Reject blanks and obvious sheet markers such as 1-5 without overfitting vendors."""
    return bool(
        value
        and len(value) >= 6
        and not any(character.isspace() for character in value)
        and re.fullmatch(r"[A-Za-z0-9._-]+", value)
    )


def cell_is_red(cell: Any) -> bool:
    colors = [getattr(cell.font, "color", None)]
    if getattr(cell.fill, "fill_type", None):
        colors.extend((getattr(cell.fill, "fgColor", None), getattr(cell.fill, "bgColor", None)))
    for color in colors:
        if color is None or getattr(color, "type", None) != "rgb":
            continue
        rgb = str(getattr(color, "rgb", "") or "").upper()
        if rgb.endswith("FF0000"):
            return True
    return False


def normalize_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                pass
    return None


def username_for(row: tuple[Any, ...]) -> tuple[str | None, bool]:
    username = clean_text(row[3].value)
    if username:
        return username, False
    email = clean_text(row[5].value)
    if email and "@" in email:
        return email.split("@", 1)[0], True
    return None, False


def select_workbook_sheet(workbook: Any) -> Any:
    if "Bookings 2026" in workbook.sheetnames:
        return workbook["Bookings 2026"]
    selected = choose_number(
        "The workbook has no sheet named 'Bookings 2026'. Select a sheet",
        list(workbook.sheetnames),
    )
    return workbook[workbook.sheetnames[selected]]


def load_sheet(path: Path) -> tuple[str, list[SheetRow]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise dwp.DWPError(
            "Spreadsheet support is not installed. Run: "
            "python3 -m pip install -r requirements-sheet.txt"
        ) from exc
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except FileNotFoundError as exc:
        raise dwp.DWPError(f"Spreadsheet not found: {path}") from exc
    except Exception as exc:
        raise dwp.DWPError(
            f"Could not read {path.name}. Use an unencrypted .xlsx or .xlsm workbook."
        ) from exc
    try:
        sheet = select_workbook_sheet(workbook)
        rows: list[SheetRow] = []
        for values in sheet.iter_rows(min_row=2, min_col=1, max_col=12):
            deployment_date = normalize_date(values[0].value, workbook.epoch)
            if deployment_date is None:
                continue
            username, from_email = username_for(values)
            rows.append(
                SheetRow(
                    row_number=values[0].row,
                    deployment_date=deployment_date,
                    username=username,
                    username_from_email=from_email,
                    new_serial=clean_text(values[9].value),
                    old_serial=clean_text(values[11].value),
                    marked_red=any(cell_is_red(cell) for cell in values),
                )
            )
        if not rows:
            raise dwp.DWPError(
                f"No dated data rows were found in columns A-L of sheet {sheet.title!r}."
            )
        return sheet.title, rows
    finally:
        workbook.close()


def latest_download() -> Path | None:
    downloads = Path.home() / "Downloads"
    candidates = [
        path
        for pattern in (f"{FILE_PREFIX}*.xlsx", f"{FILE_PREFIX}*.xlsm")
        for path in downloads.glob(pattern)
        if not path.name.startswith("~$")
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def choose_file(argument: str | None) -> Path:
    if argument:
        return Path(argument).expanduser().resolve()
    latest = latest_download()
    if latest:
        raw = input(f"Spreadsheet [{latest}]: ").strip()
        return Path(raw).expanduser().resolve() if raw else latest.resolve()
    raw = input("Spreadsheet path: ").strip()
    if not raw:
        raise dwp.DWPError(
            f"No {FILE_PREFIX!r} workbook was found in Downloads and no file was entered."
        )
    return Path(raw).expanduser().resolve()


def choose_number(label: str, options: list[str]) -> int:
    if not options:
        raise dwp.DWPError(f"No choices are available for {label}.")
    print(f"\n{label}")
    for index, option in enumerate(options, 1):
        print(f"  {index}. {option}")
    while True:
        raw = input(f"Choose 1-{len(options)}: ").strip()
        if raw.isdigit() and 1 <= int(raw) <= len(options):
            return int(raw) - 1
        print("Enter one of the listed numbers.")


def yes_no(label: str, *, default: bool = False) -> bool:
    suffix = "[Y/n]" if default else "[y/N]"
    while True:
        value = input(f"{label} {suffix}: ").strip().casefold()
        if not value:
            return default
        if value in ("y", "yes"):
            return True
        if value in ("n", "no"):
            return False
        print("Enter y or n.")


def parse_number_selection(raw: str, maximum: int) -> set[int]:
    selected: set[int] = set()
    if not raw.strip():
        return selected
    for part in raw.split(","):
        token = part.strip()
        match = re.fullmatch(r"(\d+)(?:-(\d+))?", token)
        if not match:
            raise ValueError(f"{token!r} is not a number or range")
        first = int(match.group(1))
        last = int(match.group(2) or first)
        if first > last or first < 1 or last > maximum:
            raise ValueError(f"{token!r} is outside 1-{maximum}")
        selected.update(range(first - 1, last))
    return selected


def eligible_counts(rows: Iterable[SheetRow]) -> tuple[int, int]:
    eligible = [
        row
        for row in rows
        if not row.marked_red and looks_like_serial(row.new_serial) and row.username
    ]
    return len(eligible), sum(looks_like_serial(row.old_serial) for row in eligible)


def build_actions(
    rows: list[SheetRow], selected_date: date, mode: str
) -> tuple[list[Action], Counter[str]]:
    actions: list[Action] = []
    ignored: Counter[str] = Counter()
    for row in rows:
        if row.deployment_date != selected_date:
            continue
        if row.marked_red:
            ignored["marked red"] += 1
            continue
        if not looks_like_serial(row.new_serial):
            ignored["no usable new serial in column J"] += 1
            continue
        if not row.username:
            ignored["no username in column D or email in column F"] += 1
            continue
        if mode in ("new", "both"):
            actions.append(
                Action("New deployments", row.row_number, row.username, row.new_serial, NEW_STOCK)
            )
        if mode in ("returns", "both") and looks_like_serial(row.old_serial):
            actions.append(
                Action("Old / pending return", row.row_number, row.username, row.old_serial, PENDING_RETURN)
            )
        elif mode in ("returns", "both"):
            ignored["no usable old serial in column L"] += 1

    serial_spellings: dict[str, str] = {}
    counts: Counter[str] = Counter()
    for action in actions:
        key = action.serial.casefold()
        counts[key] += 1
        serial_spellings.setdefault(key, action.serial)
    duplicates = [serial_spellings[key] for key, count in counts.items() if count > 1]
    if duplicates:
        raise dwp.DWPError(
            "The selected work contains duplicate serial numbers: " + ", ".join(duplicates)
        )
    return actions, ignored


def override_new_statuses(actions: list[Action]) -> list[Action]:
    new_indexes = [index for index, action in enumerate(actions) if action.group == "New deployments"]
    if not new_indexes:
        return actions
    print("\nNew deployments (default: Deployed - New Stock)")
    for display_index, action_index in enumerate(new_indexes, 1):
        action = actions[action_index]
        print(f"  {display_index}. row {action.row_number}: {action.serial} → {action.username}")
    while True:
        raw = input(
            "Numbers to change to Deployed - Existing Stock "
            "(comma-separated; ranges such as 2-4; Enter for none): "
        )
        try:
            selected = parse_number_selection(raw, len(new_indexes))
            break
        except ValueError as exc:
            print(f"Please correct the selection: {exc}.")
    updated = list(actions)
    for selected_index in selected:
        action_index = new_indexes[selected_index]
        updated[action_index] = replace(updated[action_index], status=EXISTING_STOCK)
    return updated


def print_preview(
    path: Path,
    sheet_name: str,
    selected_date: date,
    actions: list[Action],
    ignored: Counter[str],
    fallback_count: int,
) -> None:
    print("\nPreview")
    print(f"  File: {path}")
    print(f"  Sheet: {sheet_name}")
    print(f"  Date: {selected_date.strftime('%A %-d %B %Y')}")
    for group in ("New deployments", "Old / pending return"):
        grouped = [action for action in actions if action.group == group]
        print(f"\n{group} ({len(grouped)})")
        if not grouped:
            print("  None")
        for index, action in enumerate(grouped, 1):
            print(
                f"  {index}. row {action.row_number} | {action.serial} | "
                f"{action.username} | {action.status}"
            )
    if ignored:
        print("\nIgnored rows")
        for reason, count in sorted(ignored.items()):
            print(f"  {count} × {reason}")
    if fallback_count:
        print(
            f"\nNote: {fallback_count} selected row(s) used the column F email prefix "
            "because column D did not contain a usable username."
        )


def execute(
    client: Any, actions: list[Action], request_for: str
) -> list[Outcome]:
    outcomes: list[Outcome] = []
    total = len(actions)
    for index, action in enumerate(actions, 1):
        print(
            f"\n[{index}/{total}] {action.group}: {action.serial} → "
            f"{action.username} ({action.status})"
        )
        try:
            result = dwp.deploy_device_to_user(
                client,
                serial=action.serial,
                request_for=request_for,
                deployed_to=action.username,
                status=action.status,
                submit=True,
            )
            outcomes.append(Outcome(action, result.request_id, result.order_id, None))
        except dwp.DWPError as exc:
            request_id = (
                exc.request_id if isinstance(exc, dwp.DeploymentExecutionError) else None
            )
            outcomes.append(Outcome(action, request_id, None, str(exc)))
            print(f"Could not deploy {action.serial}: {exc}")
    return outcomes


def print_results(outcomes: list[Outcome]) -> None:
    print("\nResults")
    for group in ("New deployments", "Old / pending return"):
        grouped = [outcome for outcome in outcomes if outcome.action.group == group]
        print(f"\n{group}")
        if not grouped:
            print("  None")
        for outcome in grouped:
            if outcome.error:
                request = (
                    f" (request {outcome.request_id})" if outcome.request_id else ""
                )
                print(f"  {outcome.action.serial}: FAILED{request} — {outcome.error}")
            else:
                order = f" (order {outcome.order_id})" if outcome.order_id else ""
                print(f"  {outcome.action.serial}: request {outcome.request_id}{order}")
    failures = sum(bool(outcome.error) for outcome in outcomes)
    print(
        f"\nCompleted: {len(outcomes) - failures} succeeded, {failures} failed, "
        f"{len(outcomes)} total."
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Workbook rules:
  The newest 'Inventory Tracking - Sydney*.xlsx' or .xlsm in Downloads is used
  when FILE is omitted. 'Bookings 2026' is selected automatically; otherwise
  you choose a sheet. A=deployment date, D=username, J=new serial, L=old serial.
  Red rows and rows without a usable column-J serial are excluded before DWP.

Modes:
  --dry-run ends after the preview with zero browser or DWP API activity.
  --simulate continues after confirmation using local SIM-REQ/SIM-ORDER IDs,
  but never opens Chrome, reads cookies, contacts DWP, or changes real data.
""",
    )
    parser.add_argument("file", nargs="?", metavar="FILE", help="Optional .xlsx/.xlsm workbook. If omitted, choose the newest matching Downloads file or enter a path.")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview only. Does not open Chrome or make any DWP API requests.",
    )
    parser.add_argument("--request-for", help="Remedy login ID for every generated request. Prompts if omitted.")
    parser.add_argument(
        "--browser-profile",
        default=dwp.DEFAULT_BROWSER_PROFILE,
        help="Dedicated installed-Chrome profile for SSO, separate from normal browsing.",
    )
    parser.add_argument("--cookie-mode", action="store_true", help="Use DWP_COOKIE instead of opening Chrome. The cookie is never saved.")
    parser.add_argument(
        "--simulate",
        action="store_true",
        help="Submit into a local simulator after the preview. Produces SIM-REQ IDs with no browser, network, or DWP changes.",
    )
    parser.add_argument("--verbose", action="store_true", help="Show safe request/status diagnostics.")
    parser.add_argument("--base", default=os.getenv("DWP_BASE", dwp.DEFAULT_BASE), help="Override DWP REST base URL (HTTPS URL ending in /rest).")
    args = parser.parse_args()
    if args.cookie_mode:
        args.browser_profile = None
    parsed_base = urllib.parse.urlparse(args.base)
    if (
        parsed_base.scheme != "https"
        or not parsed_base.netloc
        or not parsed_base.path.rstrip("/").endswith("/rest")
    ):
        raise dwp.DWPError("--base must be an HTTPS DWP REST URL ending in /rest")

    path = choose_file(args.file)
    sheet_name, rows = load_sheet(path)
    dates = sorted({row.deployment_date for row in rows}, reverse=True)
    date_options = []
    for candidate in dates:
        new_count, return_count = eligible_counts(
            row for row in rows if row.deployment_date == candidate
        )
        date_options.append(
            f"{candidate.strftime('%A %-d %B %Y')} "
            f"({new_count} new, {return_count} returns)"
        )
    selected_date = dates[choose_number("Deployment date", date_options)]
    mode_index = choose_number(
        "What should be deployed?",
        ["New serials from column J", "Returns from column L", "Both new serials and returns"],
    )
    mode = ("new", "returns", "both")[mode_index]
    actions, ignored = build_actions(rows, selected_date, mode)
    if not actions:
        raise dwp.DWPError("No deployable rows remain for that date and selection.")
    actions = override_new_statuses(actions)
    selected_rows = {
        action.row_number for action in actions
    }
    fallback_count = sum(
        row.username_from_email for row in rows if row.row_number in selected_rows
    )
    print_preview(path, sheet_name, selected_date, actions, ignored, fallback_count)

    if args.dry_run:
        print("\nDry run complete. No browser was opened and no DWP API requests were made.")
        return 0
    request_for = (args.request_for or "").strip()
    if not request_for:
        suggested = getpass.getuser()
        request_for = input(f"\nRequest-for login ID [{suggested}]: ").strip() or suggested
    if not request_for or any(character.isspace() for character in request_for):
        raise dwp.DWPError("The request-for login ID cannot be empty or contain whitespace.")
    if not yes_no(f"Submit all {len(actions)} requests now?"):
        print("Cancelled before authentication. No DWP API requests were made.")
        return 0

    client = dwp.open_client(
        base=args.base,
        browser_profile=args.browser_profile,
        simulate=args.simulate,
        verbose=args.verbose,
    )
    outcomes = execute(client, actions, request_for)
    print_results(outcomes)
    return 1 if any(outcome.error for outcome in outcomes) else 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        raise SystemExit(130)
    except EOFError:
        print("\nInput ended before the import was complete.", file=sys.stderr)
        raise SystemExit(2)
    except dwp.DWPError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(2)
    except Exception:
        print(
            "Error: An unexpected problem occurred. Re-run with --verbose and "
            "report the last step shown.",
            file=sys.stderr,
        )
        raise SystemExit(2)
