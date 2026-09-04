"""Pure request and spreadsheet models used by the AutoEUDM web interface."""

from __future__ import annotations

import base64
import binascii
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO
import re
from typing import Any, Callable
import uuid

from . import eudm_inventory_import as inventory
from . import eudm_request as eudm
from .fast_workbook import FastWorkbook, FastWorkbookError
from .identifiers import is_login_id, is_serial


USER_STATUSES: tuple[tuple[str, str], ...] = (
    ("Deployed - Existing Stock", "Deployed - Existing Stock"),
    ("Deployed - New Stock", "Deployed - New Stock"),
    ("Deployed - Loan", "Loan"),
    ("Deployed - Pending Return", "Pending Return"),
    ("Deployed - Unmanaged", "Unmanaged"),
)

LOCATION_STATUSES: tuple[tuple[str, str], ...] = (
    ("Donated", "Donated"),
    ("Hold", "Hold"),
    ("New Stock", "New Stock"),
    ("Loan Stock", "Loan Stock"),
    ("Used Stock", "Used Stock"),
    ("Pending Decom", "Pending Decom"),
    ("Pending Disposal", "Pending Disposal"),
    ("Pending Pickup", "Pending Pickup"),
    ("Pending Repair", "Pending Repair"),
    ("Pending Rebuild", "Pending Rebuild"),
    ("Under Repair", "Under Repair"),
    ("Vendor Collected", "Vendor Collected"),
    ("Stolen/Lost", "Stolen/Lost"),
)

CITIES: tuple[str, ...] = (
    "Bangkok, TH",
    "Beijing, CN",
    "Brisbane, AU",
    "Calgary, CA",
    "Chicago, US",
    "Dubai, AE",
    "Dublin, IE",
    "Frankfurt, DE",
    "Geneva, CH",
    "Gurugram, IN",
    "Hong Kong, CN",
    "Houston, US",
    "Hyderabad, IN",
    "Jacksonville, US",
    "Jakarta, ID",
    "Kansas City, US",
    "Kuala Lumpur, MY",
    "London, UK",
    "Los Angeles, US",
    "Luxembourg, LU",
    "Manila, PH",
    "Melbourne, AU",
    "Mexico City, MX",
    "Minneapolis, US",
    "Mumbai, IN",
    "New York, US",
    "Paris, FR",
    "Perth, AU",
    "Philadelphia, US",
    "San Diego, US",
    "San Jose, US",
    "Santiago, Chile",
    "Sao Paulo, BR",
    "Seoul, KR",
    "Shanghai, CN",
    "Singapore, SG",
    "Sydney, AU",
    "Taipei, TW",
    "Tokyo, JP",
    "Toronto, CA",
)

MAX_WORKBOOK_BYTES = 100 * 1024 * 1024


def clean(value: Any) -> str:
    return str(value or "").strip()


def parse_serials(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = [clean(item) for item in value]
    else:
        raw = [
            item.strip()
            for item in re.split(r"[\s,;]+", clean(value))
        ]
    return [item for item in raw if item]


@dataclass(frozen=True)
class Location:
    city: str
    building: str
    floor: str
    room: str
    cabinet: str | None = None

    @classmethod
    def from_json(cls, raw: Any) -> "Location":
        source = raw if isinstance(raw, dict) else {}
        return cls(
            city=clean(source.get("city")),
            building=clean(source.get("building")),
            floor=clean(source.get("floor")),
            room=clean(source.get("room")),
            cabinet=clean(source.get("cabinet")) or None,
        )

    def display(self) -> str:
        return " → ".join(
            value
            for value in (
                self.city,
                self.building,
                self.floor,
                self.room,
                self.cabinet,
            )
            if value
        )

    def to_json(self) -> dict[str, Any]:
        return {
            "city": self.city,
            "building": self.building,
            "floor": self.floor,
            "room": self.room,
            "cabinet": self.cabinet or "",
            "display": self.display(),
        }


@dataclass(frozen=True)
class RequestSpec:
    client_id: str
    kind: str
    serials: tuple[str, ...]
    status: str
    user: str | None
    location: Location | None
    group: str
    returning_requested: bool = False
    returning_user: str | None = None
    returning_user_info: dict[str, Any] | None = None
    user_info: dict[str, Any] | None = None
    source: str | None = None
    device_allocation: str | None = None
    first_name: str | None = None
    last_name: str | None = None

    @classmethod
    def from_json(cls, raw: Any) -> "RequestSpec":
        if not isinstance(raw, dict):
            raise eudm.EUDMError("Each queue entry must be an object.")
        kind = clean(raw.get("kind"))
        returning = raw.get("returning", False)
        if not isinstance(returning, bool):
            raise eudm.EUDMError("The returning-user toggle must be true or false.")
        returning_user = clean(raw.get("returning_user")) or None
        raw_returning_info = raw.get("returning_user_info")
        if raw_returning_info is not None and not isinstance(raw_returning_info, dict):
            raise eudm.EUDMError("Returning-user details must be an object.")
        returning_user_info = None
        if isinstance(raw_returning_info, dict):
            raw_columns = raw_returning_info.get("columns", [])
            if not isinstance(raw_columns, list):
                raise eudm.EUDMError("Returning-user detail columns must be a list.")
            returning_user_info = {
                "login": clean(raw_returning_info.get("login")),
                "columns": [clean(value) for value in raw_columns if clean(value)],
            }
        raw_user_info = raw.get("user_info")
        if raw_user_info is not None and not isinstance(raw_user_info, dict):
            raise eudm.EUDMError("User details must be an object.")
        user_info = None
        if isinstance(raw_user_info, dict):
            raw_columns = raw_user_info.get("columns", [])
            if not isinstance(raw_columns, list):
                raise eudm.EUDMError("User detail columns must be a list.")
            user_info = {
                "login": clean(raw_user_info.get("login")),
                "columns": [clean(value) for value in raw_columns if clean(value)],
            }
        location = (
            Location.from_json(raw.get("location"))
            if kind in {"location", "bulk_location"}
            else None
        )
        return cls(
            client_id=clean(raw.get("id")) or uuid.uuid4().hex,
            kind=kind,
            serials=tuple(parse_serials(raw.get("serials"))),
            status=clean(raw.get("status")),
            user=clean(raw.get("user")) or None,
            location=location,
            group=clean(raw.get("group")) or "Requests",
            returning_requested=returning or bool(returning_user),
            returning_user=returning_user,
            returning_user_info=returning_user_info,
            user_info=user_info,
            source=clean(raw.get("source")) or None,
            device_allocation=clean(raw.get("device_allocation")) or None,
            first_name=clean(raw.get("first_name")) or None,
            last_name=clean(raw.get("last_name")) or None,
        )

    def validate(
        self,
        *,
        user_statuses: set[str] | None = None,
        location_statuses: set[str] | None = None,
    ) -> list[str]:
        errors: list[str] = []
        if self.kind not in {"user", "location", "bulk_location"}:
            return ["Choose Deploy to user, Add to location stock, or Bulk add to location stock."]
        if not self.serials:
            errors.append(
                "Enter one or more serial numbers."
                if self.kind == "bulk_location"
                else "Enter a serial number."
            )
        if self.kind != "bulk_location" and len(self.serials) > 1:
            errors.append("This request must contain exactly one serial number.")
        duplicate_serials = [
            serial
            for serial, count in Counter(
                item.casefold() for item in self.serials
            ).items()
            if count > 1
        ]
        if duplicate_serials:
            errors.append("Remove duplicate serial numbers from this request.")
        if any(not is_serial(serial) for serial in self.serials):
            errors.append(
                "Serial numbers must be at least 6 characters and contain only letters, numbers, periods, underscores, or hyphens."
            )

        if self.kind == "user":
            allowed = user_statuses or {value for _, value in USER_STATUSES}
            if self.status not in allowed:
                errors.append("Choose a valid status for Deploy to user.")
            if not self.user:
                errors.append("Choose the receiving user.")
            elif not is_login_id(self.user):
                errors.append("The receiving user must be a login ID, not a display name or email address.")
            if self.location:
                errors.append("Deploy to user cannot include a location.")
            if self.returning_requested:
                errors.append("Deploy to user cannot have a returning user.")
        else:
            allowed = location_statuses or {
                value for _, value in LOCATION_STATUSES
            }
            if self.status not in allowed:
                errors.append("Choose a valid status for Add to location stock.")
            if self.user:
                errors.append("Add to location stock cannot include a deployed-to user.")
            if not self.location:
                errors.append("Choose a location.")
            elif not all(
                (
                    self.location.city,
                    self.location.building,
                    self.location.floor,
                    self.location.room,
                )
            ):
                errors.append("Choose both the city and the location.")
            if self.returning_requested and not self.returning_user:
                errors.append("Choose the returning user or turn off the return option.")
            elif self.returning_user and not is_login_id(self.returning_user):
                errors.append(
                    "The returning user must be a login ID, not a display name or email address."
                )
            if self.returning_user and not self.returning_user_info:
                errors.append(
                    "Search and verify the returning user's details before submitting; "
                    "an email will be sent to them."
                )
            elif (
                self.returning_user
                and clean(self.returning_user_info.get("login")).casefold()
                != self.returning_user.casefold()
            ):
                errors.append("Verify the returning user again because the saved user details do not match.")
            if self.kind == "bulk_location" and self.returning_requested:
                errors.append("Bulk add to location stock cannot include a returning user.")
        return errors

    def device_count(self) -> int:
        return len(self.serials)

    def destination(self) -> str:
        if self.kind == "user":
            return self.user or ""
        return self.location.display() if self.location else ""

    def to_json(self) -> dict[str, Any]:
        return {
            "id": self.client_id,
            "kind": self.kind,
            "serials": list(self.serials),
            "status": self.status,
            "user": self.user or "",
            "returning": self.returning_requested,
            "returning_user": self.returning_user or "",
            "user_info": self.user_info or None,
            "returning_user_info": self.returning_user_info or None,
            "location": self.location.to_json() if self.location else None,
            "group": self.group,
            "source": self.source or "",
            "device_allocation": self.device_allocation or "",
            "first_name": self.first_name or "",
            "last_name": self.last_name or "",
            "errors": self.validate(),
            "destination": self.destination(),
            "device_count": self.device_count(),
        }


def validate_queue(
    specs: list[RequestSpec],
    request_for: str,
    *,
    user_statuses: set[str] | None = None,
    location_statuses: set[str] | None = None,
) -> dict[str, list[str]]:
    errors: dict[str, list[str]] = {}

    def add_error(key: str, message: str) -> None:
        messages = errors.setdefault(key, [])
        if message not in messages:
            messages.append(message)

    for spec in specs:
        spec_errors = spec.validate(
            user_statuses=user_statuses,
            location_statuses=location_statuses,
        )
        for message in spec_errors:
            add_error(spec.client_id, message)

    requester = clean(request_for)
    if not requester:
        add_error("_queue", "Set EUDM_REQUEST_FOR or enter a requesting login ID.")
    elif not is_login_id(requester):
        add_error(
            "_queue",
            "The requesting user must be a login ID, not a display name or email address.",
        )
    if not specs:
        add_error("_queue", "Add at least one request.")

    client_id_counts = Counter(spec.client_id for spec in specs)
    if any(count > 1 for count in client_id_counts.values()):
        add_error(
            "_queue",
            "The request queue contains duplicate internal IDs. Reopen AutoEUDM and rebuild the queue.",
        )
    if "_queue" in client_id_counts:
        add_error(
            "_queue",
            "A request used a reserved internal ID. Reopen AutoEUDM and rebuild the queue.",
        )

    owners: dict[str, list[str]] = {}
    spellings: dict[str, str] = {}
    for spec in specs:
        seen_in_request: set[str] = set()
        for serial in spec.serials:
            key = serial.casefold()
            if key in seen_in_request:
                continue
            seen_in_request.add(key)
            owners.setdefault(key, []).append(spec.client_id)
            spellings.setdefault(key, serial)
    for key, client_ids in owners.items():
        if len(client_ids) > 1:
            message = (
                f"Serial {spellings[key]} appears in more than one queued request."
            )
            for client_id in dict.fromkeys(client_ids):
                add_error(client_id, message)
    return errors


@dataclass
class WorkbookImport:
    import_id: str
    filename: str
    sheets: dict[str, list[inventory.SheetRow]]
    _summary_cache: dict[str, Any] | None = field(default=None, repr=False)
    _inspection_cache: dict[str, Any] | None = field(default=None, repr=False)

    @staticmethod
    def decode_upload(filename: str, encoded: str) -> bytes:
        """Validate and decode one browser workbook upload exactly once."""
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise eudm.EUDMError("Choose an .xlsx or .xlsm workbook.")
        try:
            payload = base64.b64decode(encoded, validate=True)
        except (binascii.Error, ValueError) as exc:
            raise eudm.EUDMError("The uploaded workbook data was invalid.") from exc
        if not payload:
            raise eudm.EUDMError("The uploaded workbook was empty.")
        if len(payload) > MAX_WORKBOOK_BYTES:
            raise eudm.EUDMError("The workbook is larger than the 100 MB local limit.")
        return payload

    @staticmethod
    def inspect_payload(filename: str, payload: bytes) -> dict[str, Any]:
        """Return selectable headings before committing to a column mapping."""
        try:
            with FastWorkbook(payload) as workbook:
                return WorkbookImport._inspect_fast_workbook(filename, workbook)
        except FastWorkbookError:
            # Keep the compatibility reader for unusual but valid workbooks.
            pass
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
        except Exception as exc:
            raise eudm.EUDMError("Could not read the workbook. Use an unencrypted .xlsx or .xlsm file.") from exc
        try:
            return WorkbookImport._inspect_openpyxl_workbook(filename, workbook)
        finally:
            workbook.close()

    @staticmethod
    def _inspect_fast_workbook(filename: str, workbook: FastWorkbook) -> dict[str, Any]:
        """Inspect the small header window without creating openpyxl cells."""
        source_sheets = (
            ["Bookings 2026"]
            if "Bookings 2026" in workbook.sheet_names
            else list(workbook.sheet_names)
        )
        sheets = []
        for sheet_name in source_sheets:
            headings: list[str] = []
            for row in workbook.iter_rows(sheet_name):
                if row.row_number > 25:
                    break
                values = {
                    column: inventory.clean_text(row.cells[column][0])
                    for column in sorted(row.cells)
                }
                if any(
                    inventory.normalized_header(value)
                    in {"date", "deployment date", "booking date"}
                    for value in values.values()
                ):
                    headings = [value for value in values.values() if value]
                    break
            if headings:
                sheets.append({"name": sheet_name, "headings": headings})
        if not sheets:
            raise eudm.EUDMError("No sheet with a Date heading was found.")
        return {
            "filename": filename,
            "default_sheet": sheets[0]["name"],
            "sheets": sheets,
            "needs_mapping": True,
        }

    @staticmethod
    def inspect_path(filename: str, path: str | Any) -> dict[str, Any]:
        """Inspect a local workbook without copying it through the browser."""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise eudm.EUDMError(
                "Could not read the workbook. Use an unencrypted .xlsx or .xlsm file."
            ) from exc
        try:
            return WorkbookImport._inspect_openpyxl_workbook(filename, workbook)
        finally:
            workbook.close()

    @staticmethod
    def _inspect_openpyxl_workbook(filename: str, workbook: Any) -> dict[str, Any]:
        """Read only the headings needed to populate the import mapping step."""
        # The normal ALM source is the current Bookings sheet. Looking there
        # first avoids opening every archival/notes tab in a large workbook.
        source_sheets = (
            [workbook["Bookings 2026"]]
            if "Bookings 2026" in workbook.sheetnames
            else list(workbook.worksheets)
        )
        sheets = []
        for sheet in source_sheets:
            headings: list[str] = []
            for row in sheet.iter_rows(
                min_row=1, max_row=min(25, int(sheet.max_row or 25))
            ):
                values = [inventory.clean_text(cell.value) for cell in row]
                # Tracking headers always include Date; only offer header-like rows.
                if any(
                    inventory.normalized_header(value)
                    in {"date", "deployment date", "booking date"}
                    for value in values
                ):
                    headings = [value for value in values if value]
                    break
            if headings:
                sheets.append({"name": sheet.title, "headings": headings})
        if not sheets:
            raise eudm.EUDMError("No sheet with a Date heading was found.")
        default_sheet = sheets[0]["name"]
        return {
            "filename": filename,
            "default_sheet": default_sheet,
            "sheets": sheets,
            "needs_mapping": True,
        }

    @staticmethod
    def _fast_header_indexes(
        workbook: FastWorkbook,
        sheet_name: str,
        columns: inventory.ImportColumns,
    ) -> tuple[int, dict[str, int], int]:
        """Find the configured headers while scanning only the first 25 rows."""
        desired = {
            "username": columns.username,
            "deployment_serial": columns.deployment_serial,
            "returned_device": columns.returned_device,
            "pending_return": columns.pending_return,
            "enabled": columns.enabled,
            "device_allocation": columns.device_allocation,
            "new_asset_status": columns.new_asset_status,
            "first_name": columns.first_name,
            "last_name": columns.last_name,
        }
        targets = {
            key: inventory.normalized_header(value)
            for key, value in desired.items()
        }
        date_titles = {"date", "deployment date", "booking date"}
        for row in workbook.iter_rows(sheet_name):
            if row.row_number > 25:
                break
            found: dict[str, int] = {}
            for column, (value, _style) in row.cells.items():
                title = inventory.normalized_header(inventory.clean_text(value))
                if title:
                    found[title] = column
            indexes = {key: found.get(title) for key, title in targets.items()}
            if not indexes["username"] or not indexes["deployment_serial"] or not indexes["pending_return"]:
                continue
            if not indexes["enabled"] and not columns.enabled:
                for title in (
                    "attend",
                    "attended",
                    "attendance",
                    "eligible",
                    "enabled",
                ):
                    if title in found:
                        indexes["enabled"] = found[title]
                        break
            date_index = next(
                (found[title] for title in date_titles if title in found),
                None,
            )
            if date_index:
                return (
                    row.row_number,
                    {key: value or 0 for key, value in indexes.items()},
                    date_index,
                )
        required = (
            columns.username,
            columns.deployment_serial,
            columns.pending_return,
            "Date",
        )
        missing = ", ".join(f"{name!r}" for name in required)
        raise eudm.EUDMError(
            "Could not find the required spreadsheet headers. Check ALM Workbook settings: "
            + missing
        )

    @staticmethod
    def _fast_cell_value(row: Any, column: int) -> Any:
        cell = row.cells.get(column)
        return cell[0] if cell else None

    @staticmethod
    def _fast_normalize_date(value: Any, epoch: datetime, from_excel: Any) -> date | None:
        """Convert fast-reader date values without importing openpyxl per row."""
        if isinstance(value, (int, float)):
            try:
                converted = from_excel(value, epoch)
            except (TypeError, ValueError, OverflowError):
                return None
            return converted.date() if isinstance(converted, datetime) else converted
        return inventory.normalize_date(value, epoch)

    @staticmethod
    def _row_to_cache(row: inventory.SheetRow) -> dict[str, Any]:
        return {
            "row_number": row.row_number,
            "deployment_date": row.deployment_date.isoformat(),
            "username": row.username,
            "deployment_serial": row.deployment_serial,
            "returned_device_serial": row.returned_device_serial,
            "pending_return_serial": row.pending_return_serial,
            "marked_red": row.marked_red,
            "enabled": row.enabled,
            "date_group": row.date_group,
            "returned_device_column_present": row.returned_device_column_present,
            "device_allocation": row.device_allocation,
            "new_asset_status": row.new_asset_status,
            "new_joiner": row.new_joiner,
            "first_name": row.first_name,
            "last_name": row.last_name,
            "deployment_status_hint": row.deployment_status_hint,
            "returned_device_status_hint": row.returned_device_status_hint,
        }

    def cache_json(self) -> dict[str, Any]:
        """Serialize the parsed rows for fast reuse of an unchanged local file."""
        return {
            "version": 2,
            "filename": self.filename,
            "sheets": {
                name: [self._row_to_cache(row) for row in rows]
                for name, rows in self.sheets.items()
            },
            "summary": self.summary(),
            "inspection": self._inspection_cache,
        }

    @classmethod
    def from_cache_json(
        cls,
        import_id: str,
        filename: str,
        raw: Any,
    ) -> "WorkbookImport | None":
        """Restore a cache entry, returning None when it is incomplete."""
        try:
            if not isinstance(raw, dict) or raw.get("version") != 2:
                return None
            raw_sheets = raw.get("sheets")
            if not isinstance(raw_sheets, dict) or not raw_sheets:
                return None
            sheets: dict[str, list[inventory.SheetRow]] = {}
            for name, raw_rows in raw_sheets.items():
                if not isinstance(name, str) or not isinstance(raw_rows, list):
                    return None
                rows: list[inventory.SheetRow] = []
                for raw_row in raw_rows:
                    if not isinstance(raw_row, dict):
                        return None
                    rows.append(
                        inventory.SheetRow(
                            row_number=int(raw_row["row_number"]),
                            deployment_date=date.fromisoformat(
                                str(raw_row["deployment_date"])
                            ),
                            username=raw_row.get("username") or None,
                            deployment_serial=raw_row.get("deployment_serial") or None,
                            returned_device_serial=raw_row.get("returned_device_serial") or None,
                            pending_return_serial=raw_row.get("pending_return_serial") or None,
                            marked_red=bool(raw_row.get("marked_red")),
                            enabled=bool(raw_row.get("enabled")),
                            date_group=int(raw_row.get("date_group", 1)),
                            returned_device_column_present=bool(
                                raw_row.get("returned_device_column_present", True)
                            ),
                            device_allocation=raw_row.get("device_allocation") or None,
                            new_asset_status=raw_row.get("new_asset_status") or None,
                            new_joiner=bool(raw_row.get("new_joiner")),
                            first_name=raw_row.get("first_name") or None,
                            last_name=raw_row.get("last_name") or None,
                            deployment_status_hint=raw_row.get("deployment_status_hint") or None,
                            returned_device_status_hint=raw_row.get("returned_device_status_hint") or None,
                        )
                    )
                sheets[name] = rows
            summary = raw.get("summary")
            inspection = raw.get("inspection")
            cached_import_id = import_id
            if isinstance(summary, dict):
                try:
                    # Keep the persisted payload ID when one exists. Drafts
                    # resumed after an application restart use this ID to
                    # restore the workbook without re-reading the source.
                    cached_import_id = uuid.UUID(
                        str(summary.get("import_id", ""))
                    ).hex
                except (ValueError, AttributeError, TypeError):
                    pass
            return cls(
                cached_import_id,
                filename,
                sheets,
                summary if isinstance(summary, dict) else None,
                inspection if isinstance(inspection, dict) else None,
            )
        except (KeyError, TypeError, ValueError, OverflowError):
            return None

    @classmethod
    def from_payload(
        cls,
        filename: str,
        payload: bytes,
        *,
        columns: inventory.ImportColumns | None = None,
        on_progress: Callable[[str, int, int], None] | None = None,
    ) -> "WorkbookImport":
        """Parse an ALM workbook, using the streaming reader when possible."""
        try:
            return cls._from_fast_payload(
                filename,
                payload,
                columns=columns,
                on_progress=on_progress,
            )
        except FastWorkbookError:
            # Some valid workbooks use XML features outside this focused
            # reader. Keep openpyxl as a compatibility path for those files.
            return cls._from_openpyxl_payload(
                filename,
                payload,
                columns=columns,
                on_progress=on_progress,
            )

    @classmethod
    def _from_fast_payload(
        cls,
        filename: str,
        payload: bytes,
        *,
        columns: inventory.ImportColumns | None = None,
        on_progress: Callable[[str, int, int], None] | None = None,
    ) -> "WorkbookImport":
        try:
            from openpyxl.utils.datetime import from_excel
        except ImportError as exc:
            raise FastWorkbookError("Spreadsheet support is not installed.") from exc
        workbook = FastWorkbook(payload)
        try:
            source_sheets = (
                ["Bookings 2026"]
                if "Bookings 2026" in workbook.sheet_names
                else list(workbook.sheet_names)
            )
            max_rows = {
                name: workbook.sheet_max_row(name)
                for name in source_sheets
            }
            total_rows = sum(max(0, max_rows[name] - 1) for name in source_sheets)
            processed_rows = 0
            selected_columns = columns or inventory.ImportColumns()
            sheets: dict[str, list[inventory.SheetRow]] = {}
            for sheet_name in source_sheets:
                try:
                    header_row, indexes, date_index = cls._fast_header_indexes(
                        workbook,
                        sheet_name,
                        selected_columns,
                    )
                except eudm.EUDMError:
                    continue
                rows: list[inventory.SheetRow] = []
                current_date = None
                current_fill = None
                date_group = 0
                for row in workbook.iter_rows(sheet_name):
                    if row.row_number <= header_row:
                        continue
                    processed_rows += 1
                    if on_progress and (
                        processed_rows == total_rows or processed_rows % 150 == 0
                    ):
                        on_progress(sheet_name, processed_rows, total_rows)
                    date_value = cls._fast_cell_value(row, date_index)
                    date_cell = row.cells.get(date_index, (None, 0))
                    explicit_date = cls._fast_normalize_date(
                        date_value,
                        workbook.epoch,
                        from_excel,
                    )
                    if explicit_date is not None:
                        fill_key = workbook.fill_key(date_cell[1])
                        if explicit_date != current_date:
                            date_group = 1
                        elif fill_key != current_fill:
                            date_group += 1
                        current_date = explicit_date
                        current_fill = fill_key
                        deployment_date = explicit_date
                    elif current_date is not None and any(
                        inventory.clean_text(value)
                        for column, (value, _style) in row.cells.items()
                        if column != date_index
                    ):
                        if date_index in row.cells:
                            fill_key = workbook.fill_key(date_cell[1])
                            if fill_key != current_fill:
                                date_group += 1
                                current_fill = fill_key
                        deployment_date = current_date
                    else:
                        continue
                    deployment_serial, deployment_status_hint = inventory.serial_and_status_hint(
                        cls._fast_cell_value(row, indexes["deployment_serial"])
                        if indexes["deployment_serial"]
                        else None
                    )
                    returned_device_serial, returned_device_status_hint = inventory.serial_and_status_hint(
                        cls._fast_cell_value(row, indexes["returned_device"])
                        if indexes["returned_device"]
                        else None,
                        returned_device=True,
                    )
                    rows.append(
                        inventory.SheetRow(
                            row_number=row.row_number,
                            deployment_date=deployment_date,
                            username=inventory.clean_text(
                                cls._fast_cell_value(row, indexes["username"])
                            ),
                            deployment_serial=deployment_serial,
                            returned_device_serial=returned_device_serial,
                            pending_return_serial=inventory.clean_text(
                                cls._fast_cell_value(row, indexes["pending_return"])
                            ) if indexes["pending_return"] else None,
                            # Font colour is presentation only. Eligibility is
                            # controlled by the configured TRUE/FALSE column.
                            marked_red=False,
                            enabled=inventory.enabled_column_allows(
                                cls._fast_cell_value(row, indexes["enabled"])
                            ) if indexes["enabled"] else True,
                            date_group=date_group,
                            returned_device_column_present=bool(
                                indexes["returned_device"]
                            ),
                            device_allocation=inventory.clean_text(
                                cls._fast_cell_value(row, indexes["device_allocation"])
                            ) if indexes["device_allocation"] else None,
                            new_asset_status=inventory.clean_text(
                                cls._fast_cell_value(row, indexes["new_asset_status"])
                            ) if indexes["new_asset_status"] else None,
                            new_joiner=inventory.row_contains_new_joiner(
                                value for value, _style in row.cells.values()
                            ),
                            first_name=inventory.clean_text(
                                cls._fast_cell_value(row, indexes["first_name"])
                            ) if indexes["first_name"] else None,
                            last_name=inventory.clean_text(
                                cls._fast_cell_value(row, indexes["last_name"])
                            ) if indexes["last_name"] else None,
                            deployment_status_hint=deployment_status_hint,
                            returned_device_status_hint=returned_device_status_hint,
                        )
                    )
                if rows:
                    sheets[sheet_name] = rows
            if on_progress:
                on_progress(
                    source_sheets[-1] if source_sheets else "Workbook",
                    processed_rows,
                    total_rows,
                )
        finally:
            workbook.close()
        if not sheets:
            raise eudm.EUDMError(
                "No dated rows were found with the configured spreadsheet headers."
            )
        return cls(uuid.uuid4().hex, filename, sheets)

    @classmethod
    def _from_openpyxl_payload(
        cls,
        filename: str,
        payload: bytes,
        *,
        columns: inventory.ImportColumns | None = None,
        on_progress: Callable[[str, int, int], None] | None = None,
    ) -> "WorkbookImport":
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(
                # The importer reads values and TRUE/FALSE eligibility only.
                # Streaming cells avoids loading a large tracking workbook's
                # full style and formula graph into memory.
                BytesIO(payload), data_only=True, read_only=True
            )
        except Exception as exc:
            raise eudm.EUDMError(
                "Could not read the workbook. Use an unencrypted .xlsx or .xlsm file."
            ) from exc
        sheets: dict[str, list[inventory.SheetRow]] = {}
        try:
            # Bookings 2026 is the established source. Avoid scanning every
            # archival/notes tab in a large tracking workbook when it exists;
            # otherwise keep every sheet available for the user to choose.
            source_sheets = (
                [workbook["Bookings 2026"]]
                if "Bookings 2026" in workbook.sheetnames
                else list(workbook.worksheets)
            )
            total_rows = 0
            for sheet in source_sheets:
                if sheet.max_row is None:
                    # Some generated workbooks omit the sheet dimension. Ask
                    # openpyxl to establish it once so progress is meaningful.
                    sheet.calculate_dimension(force=True)
                total_rows += max(0, int(sheet.max_row or 1) - 1)
            processed_rows = 0
            selected_columns = columns or inventory.ImportColumns()
            for sheet in source_sheets:
                rows: list[inventory.SheetRow] = []
                try:
                    header_row, indexes, date_index = inventory.find_column_indexes(sheet, selected_columns)
                except eudm.EUDMError:
                    continue
                # The new-joiner marker can appear in any column, so keep the
                # full reported worksheet width while reading rows. Only the
                # resulting boolean is retained in each SheetRow.
                max_column = max(int(sheet.max_column or 1), date_index, *indexes.values())
                current_date = None
                current_fill = None
                date_group = 0
                for row_offset, values in enumerate(
                    sheet.iter_rows(min_row=header_row + 1, min_col=1, max_col=max_column),
                    start=1,
                ):
                    processed_rows += 1
                    if on_progress and (
                        processed_rows == total_rows or processed_rows % 150 == 0
                    ):
                        on_progress(sheet.title, processed_rows, total_rows)
                    row_number = getattr(values[0], "row", header_row + row_offset)
                    date_cell = values[date_index - 1]
                    explicit_date = inventory.normalize_date(
                        date_cell.value, workbook.epoch
                    )
                    if explicit_date is not None:
                        fill_key = inventory.background_fill_key(date_cell)
                        if explicit_date != current_date:
                            date_group = 1
                        elif fill_key != current_fill:
                            date_group += 1
                        current_date = explicit_date
                        current_fill = fill_key
                        deployment_date = explicit_date
                    elif current_date is not None and any(
                        inventory.clean_text(cell.value)
                        for index, cell in enumerate(values)
                        if index != date_index - 1
                    ):
                        # Merged/continued date cells are blank after the
                        # first row; retain their date section and detect a
                        # new Col A fill section when the styled cell changes.
                        fill_key = inventory.background_fill_key(date_cell)
                        if getattr(date_cell, "has_style", False) and fill_key != current_fill:
                            date_group += 1
                            current_fill = fill_key
                        deployment_date = current_date
                    else:
                        continue
                    deployment_serial, deployment_status_hint = inventory.serial_and_status_hint(
                        values[indexes["deployment_serial"] - 1].value
                        if indexes["deployment_serial"]
                        else None
                    )
                    returned_device_serial, returned_device_status_hint = inventory.serial_and_status_hint(
                        values[indexes["returned_device"] - 1].value
                        if indexes["returned_device"]
                        else None,
                        returned_device=True,
                    )
                    rows.append(
                        inventory.SheetRow(
                            row_number=row_number,
                            deployment_date=deployment_date,
                            username=inventory.username_for(values, indexes["username"]),
                            deployment_serial=deployment_serial,
                            returned_device_serial=returned_device_serial,
                            pending_return_serial=inventory.clean_text(values[indexes["pending_return"] - 1].value) if indexes["pending_return"] else None,
                            # Font colour is presentation only. Eligibility is
                            # controlled by the configured TRUE/FALSE column.
                            marked_red=False,
                            enabled=inventory.enabled_column_allows(
                                values[indexes["enabled"] - 1].value
                            ) if indexes["enabled"] else True,
                            date_group=date_group,
                            returned_device_column_present=bool(indexes["returned_device"]),
                            device_allocation=inventory.clean_text(values[indexes["device_allocation"] - 1].value) if indexes["device_allocation"] else None,
                            new_asset_status=inventory.clean_text(values[indexes["new_asset_status"] - 1].value) if indexes["new_asset_status"] else None,
                            new_joiner=inventory.row_contains_new_joiner(values),
                            first_name=inventory.clean_text(values[indexes["first_name"] - 1].value) if indexes["first_name"] else None,
                            last_name=inventory.clean_text(values[indexes["last_name"] - 1].value) if indexes["last_name"] else None,
                            deployment_status_hint=deployment_status_hint,
                            returned_device_status_hint=returned_device_status_hint,
                        )
                    )
                if rows:
                    sheets[sheet.title] = rows
            if on_progress:
                on_progress(
                    source_sheets[-1].title if source_sheets else "Workbook",
                    processed_rows,
                    total_rows,
                )
        finally:
            workbook.close()
        if not sheets:
            raise eudm.EUDMError(
                "No dated rows were found with the configured spreadsheet headers."
            )
        return cls(uuid.uuid4().hex, filename, sheets)

    def summary(self) -> dict[str, Any]:
        if self._summary_cache is not None:
            return {
                **self._summary_cache,
                "import_id": self.import_id,
                "filename": self.filename,
            }
        sheet_summaries = []
        for name, rows in self.sheets.items():
            rows_by_date: dict[date, list[inventory.SheetRow]] = {}
            for row in rows:
                rows_by_date.setdefault(row.deployment_date, []).append(row)
            dates = []
            for selected in sorted(rows_by_date, reverse=True):
                selected_rows = rows_by_date[selected]
                deployment_count, returned_device_count, pending_return_count = inventory.eligible_counts(
                    selected_rows
                )
                missing_username_deployment_count = sum(
                    row.enabled
                    and inventory.looks_like_serial(row.deployment_serial)
                    and not str(row.username or "").strip()
                    for row in selected_rows
                )
                groups = []
                for group_number in sorted({row.date_group for row in selected_rows}):
                    group_rows = [row for row in selected_rows if row.date_group == group_number]
                    group_deployments, group_returned, group_pending = inventory.eligible_counts(
                        group_rows
                    )
                    group_missing_username_deployments = sum(
                        row.enabled
                        and inventory.looks_like_serial(row.deployment_serial)
                        and not str(row.username or "").strip()
                        for row in group_rows
                    )
                    groups.append(
                        {
                            "value": str(group_number),
                            "row_count": len(group_rows),
                            "eligible_row_count": len(
                                inventory.eligible_rows(
                                    group_rows,
                                )
                            ),
                            "deployment_count": group_deployments,
                            "missing_username_deployment_count": group_missing_username_deployments,
                            "returned_device_count": group_returned,
                            "pending_return_count": group_pending,
                        }
                    )
                warning_rows = inventory.attended_rows_missing_return_serials(selected_rows)
                dates.append(
                    {
                        "value": selected.isoformat(),
                        "label": inventory.format_date_label(selected),
                        "deployment_count": deployment_count,
                        "missing_username_deployment_count": missing_username_deployment_count,
                        "returned_device_count": returned_device_count,
                        "pending_return_count": pending_return_count,
                        "row_count": len(selected_rows),
                        "eligible_row_count": len(
                            inventory.eligible_rows(selected_rows)
                        ),
                        "groups": groups,
                        "warnings": [
                            {
                                "row_number": row.row_number,
                                "username": row.username or "",
                                "missing_returned": (
                                    row.returned_device_column_present
                                    and not inventory.looks_like_serial(row.returned_device_serial)
                                ),
                                "missing_pending": not inventory.looks_like_serial(row.pending_return_serial),
                            }
                            for row in warning_rows
                        ],
                    }
                )
            sheet_summaries.append({"name": name, "dates": dates})
        default_sheet = (
            "Bookings 2026"
            if "Bookings 2026" in self.sheets
            else next(iter(self.sheets))
        )
        return {
            "import_id": self.import_id,
            "filename": self.filename,
            "default_sheet": default_sheet,
            "sheets": sheet_summaries,
        }

    def prepare(
        self,
        sheet_name: str,
        selected_date: str | list[str],
        mode: str,
        location: Location | None = None,
        group_selection: str | int | None = None,
        group_selections: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        if sheet_name not in self.sheets:
            raise eudm.EUDMError("Choose one of the workbook's available sheets.")
        selected_modes = {part.strip() for part in mode.split(",") if part.strip()}
        allowed_modes = {"deployments", "returned_devices", "pending_returns"}
        if not selected_modes or not selected_modes.issubset(allowed_modes | {"all"}):
            raise eudm.EUDMError("Choose what to import.")
        if "all" in selected_modes:
            selected_modes = allowed_modes
        raw_dates = selected_date if isinstance(selected_date, list) else [selected_date]
        if not raw_dates:
            raise eudm.EUDMError("Choose at least one deployment date.")
        chosen_dates: list[date] = []
        for raw_date in raw_dates:
            try:
                chosen = date.fromisoformat(str(raw_date).strip())
            except ValueError as exc:
                raise eudm.EUDMError("Choose a valid deployment date.") from exc
            if chosen not in chosen_dates:
                chosen_dates.append(chosen)
        selections = group_selections if isinstance(group_selections, dict) else {}
        if isinstance(group_selection, dict):
            selections = group_selection
            group_selection = None
        actions: list[inventory.Action] = []
        ignored: Counter[str] = Counter()
        missing_username_deployments: list[dict[str, Any]] = []
        rows_by_date: dict[date, list[inventory.SheetRow]] = {}
        for row in self.sheets[sheet_name]:
            rows_by_date.setdefault(row.deployment_date, []).append(row)
        for chosen_date in chosen_dates:
            selected_group = selections.get(chosen_date.isoformat(), group_selection)
            selected_rows = rows_by_date.get(chosen_date, [])
            date_actions, date_ignored = inventory.build_actions(
                selected_rows,
                chosen_date,
                ",".join(sorted(selected_modes)),
                group_selection=selected_group,
            )
            actions.extend(date_actions)
            ignored.update(date_ignored)
            if "deployments" in selected_modes:
                raw_group = str(selected_group or "all").strip().casefold()
                selected_group_number = None if raw_group in {"", "all"} else int(raw_group)
                for row in selected_rows:
                    if selected_group_number is not None and row.date_group != selected_group_number:
                        continue
                    if not row.enabled or str(row.username or "").strip():
                        continue
                    if not inventory.looks_like_serial(row.deployment_serial):
                        continue
                    missing_username_deployments.append(
                        {
                            "row_number": row.row_number,
                            "date": chosen_date.isoformat(),
                            "serial": str(row.deployment_serial).strip(),
                        }
                    )
        if not actions and not missing_username_deployments:
            raise eudm.EUDMError(
                "No deployable rows remain for the selected dates and sections."
            )
        serial_spellings: dict[str, str] = {}
        occurrences: dict[str, list[inventory.Action]] = {}
        for action in actions:
            key = action.serial.casefold()
            serial_spellings.setdefault(key, action.serial)
            occurrences.setdefault(key, []).append(action)
        duplicate_keys = [
            key for key, matching in occurrences.items() if len(matching) > 1
        ]
        if duplicate_keys:
            duplicate_details = []
            for key in duplicate_keys[:12]:
                locations = "; ".join(
                    f"{action.group} row {action.row_number} ({action.username})"
                    for action in occurrences[key]
                )
                duplicate_details.append(f"{serial_spellings[key]} — {locations}")
            extra = len(duplicate_keys) - len(duplicate_details)
            if extra:
                duplicate_details.append(
                    f"and {extra} more duplicate serial{'' if extra == 1 else 's'}"
                )
            scope = "selected date/section" if len(chosen_dates) == 1 else "selected dates/sections"
            raise eudm.EUDMError(
                f"Duplicate serial numbers were found across the {scope}. "
                "Each serial can only be imported once; correct the workbook or "
                "change the date or section selection before continuing: "
                + "; ".join(duplicate_details)
            )
        requests = []
        for action in actions:
            cleaned_serial, _ = inventory.serial_and_status_hint(
                action.serial,
                returned_device=action.group == "Returned devices",
            )
            if action.kind == "location" and (
                not location
                or not all(
                    (
                        location.city,
                        location.building,
                        location.floor,
                        location.room,
                    )
                )
            ):
                raise eudm.EUDMError("Choose a complete destination for returned devices.")
            request = RequestSpec(
                client_id=uuid.uuid4().hex,
                kind=action.kind,
                # Strip recognised ALM status suffixes again at this boundary.
                # This also protects resumed drafts created by an older parser.
                serials=(cleaned_serial or action.serial,),
                status=(
                    action.status
                    if action.group == "Pending returns" or action.status_preselected
                    else ""
                ),
                user=action.username if action.kind == "user" else None,
                location=location if action.kind == "location" else None,
                group=action.group,
                returning_requested=action.group == "Returned devices",
                returning_user=(
                    action.username if action.group == "Returned devices" else None
                ),
                source=f"{self.filename} · {sheet_name}",
                device_allocation=action.device_allocation,
                first_name=action.first_name,
                last_name=action.last_name,
            ).to_json()
            request["new_asset_status"] = action.new_asset_status or ""
            request["has_returned_device_serial"] = action.has_returned_device_serial
            request["has_pending_return_serial"] = action.has_pending_return_serial
            request["new_joiner"] = action.new_joiner
            requests.append(request)
        requests.sort(
            key=lambda request: 0
            if request["group"] == "Deployments" else 1 if request["group"] == "Returned devices" else 2
        )
        return {
            "requests": requests,
            "warnings": {
                "missing_username_deployments": missing_username_deployments,
            },
            "dates": [chosen.isoformat() for chosen in chosen_dates],
            "ignored": [
                {"reason": reason, "count": count}
                for reason, count in sorted(ignored.items())
            ],
            "counts": {
                "requests": len(requests),
                "deployments": sum(
                    request["group"] == "Deployments"
                    for request in requests
                ),
                "returned_devices": sum(
                    request["group"] == "Returned devices"
                    for request in requests
                ),
                "pending_returns": sum(
                    request["group"] == "Pending returns"
                    for request in requests
                ),
            },
        }

    @staticmethod
    def backlog_key(serial: str, username: str) -> str:
        return "\u0000".join(
            " ".join(str(value or "").split()).casefold()
            for value in (serial, username)
        )

    def prepare_backlog(
        self,
        sheet_name: str,
        days_back: int,
        include_today: bool,
        ignored_keys: set[str] | None = None,
        *,
        today: date | None = None,
    ) -> dict[str, Any]:
        if sheet_name not in self.sheets:
            raise eudm.EUDMError("Choose one of the workbook's available sheets.")
        try:
            days = int(days_back)
        except (TypeError, ValueError) as exc:
            raise eudm.EUDMError("Choose how many days of backlog to check.") from exc
        if not 1 <= days <= 3650:
            raise eudm.EUDMError("Backlog range must be between 1 and 3650 days.")
        current_day = today or date.today()
        start_day = current_day - timedelta(days=days)
        ignored = ignored_keys or set()
        filtered: Counter[str] = Counter()
        username_occurrences: dict[int, int] = {}
        username_totals: dict[str, int] = {}
        for row_index, row in enumerate(self.sheets[sheet_name]):
            if not " ".join(str(row.username or "").split()):
                continue
            username_key = " ".join(str(row.username or "").split()).casefold()
            occurrence = username_totals.get(username_key, 0) + 1
            username_occurrences[row_index] = occurrence
            username_totals[username_key] = occurrence
        candidates: list[dict[str, Any]] = []
        ignored_count = 0
        for row_index, row in enumerate(self.sheets[sheet_name]):
            username_key = " ".join(str(row.username or "").split()).casefold()
            if row.deployment_date < start_day or row.deployment_date > current_day:
                filtered["outside_range"] += 1
                continue
            if row.deployment_date == current_day and not include_today:
                filtered["today_excluded"] += 1
                continue
            if not inventory.looks_like_serial(row.deployment_serial):
                filtered["missing_serial"] += 1
                continue
            if not " ".join(str(row.username or "").split()):
                filtered["missing_username"] += 1
                continue
            current_status = " ".join(str(row.new_asset_status or "").split())
            # The workbook's status is authoritative here: backlog mode is
            # specifically for rows whose *New Asset Status* is not the
            # completed "Deployed" value.  Do not discard other status text
            # merely because it begins with that word; ALM uses longer status
            # labels which still need an explicit review in this workflow.
            if current_status.casefold() == "deployed":
                filtered["already_deployed"] += 1
                continue
            serial = str(row.deployment_serial).strip()
            username = str(row.username).strip()
            key = self.backlog_key(serial, username)
            if key in ignored:
                ignored_count += 1
                filtered["ignored"] += 1
                continue
            occurrence = username_occurrences[row_index]
            candidates.append({
                # The row number makes repeated username/serial combinations
                # distinct review items instead of silently collapsing them.
                "id": f"{self.import_id}-{row.row_number}-{occurrence}",
                "row_number": row.row_number,
                "date": row.deployment_date.isoformat(),
                "serial": serial,
                "username": username,
                "first_name": row.first_name or "",
                "last_name": row.last_name or "",
                "username_occurrence": occurrence,
                "username_occurrence_total": username_totals[username_key],
                "current_status": current_status or "No status",
                "device_allocation": row.device_allocation or "",
                "new_joiner": row.new_joiner,
                "attending": row.enabled,
                "included": row.enabled,
                "default_excluded": not row.enabled,
                "backlog_ignored": False,
                "status": row.deployment_status_hint or "",
            })
        return {
            "mode": "backlog",
            "filename": self.filename,
            "sheet": sheet_name,
            "days_back": days,
            "include_today": bool(include_today),
            "start_date": start_day.isoformat(),
            "end_date": current_day.isoformat(),
            "candidates": candidates,
            "ignored_count": ignored_count,
            "counts": {
                "candidates": len(candidates),
                "already_deployed": filtered["already_deployed"],
                "ignored": filtered["ignored"],
                "outside_range": filtered["outside_range"],
                "today_excluded": filtered["today_excluded"],
                "missing_serial": filtered["missing_serial"],
                "missing_username": filtered["missing_username"],
            },
        }
