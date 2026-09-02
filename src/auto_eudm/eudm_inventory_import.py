#!/usr/bin/env python3
"""Guided importer for Sydney inventory tracking workbooks.

The importer reads a dated sheet, previews the resulting device changes, then
either stops (--dry-run), submits real EUDM requests, or submits fully local
simulation requests (--simulate). It never submits before showing the preview
and receiving a final confirmation.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
import re
import sys
from typing import Any, Iterable

from .bootstrap import browser_runtime_required, ensure_runtime
from . import eudm_request as eudm
from .cli_common import (
    add_runtime_arguments,
    console,
    open_client,
    request_for as resolve_request_for,
    start_run,
    validate_runtime_args,
)
from .eudm_config import AppConfig
from .identifiers import is_login_id, is_serial
from .user_assignments import (
    DeploymentOutcome,
    UserDeployment,
    UserDeploymentRunner,
    print_grouped_results,
)


NEW_STOCK = "Deployed - New Stock"
EXISTING_STOCK = "Deployed - Existing Stock"
PENDING_RETURN = "Pending Return"
FILE_PREFIX = "Inventory Tracking - Sydney"
CLI_GROUPS = ("Deployments", "Pending returns")


@dataclass(frozen=True)
class ImportColumns:
    """Header titles used by the tracking workbook, independent of column order."""

    username: str = "Username"
    deployment_serial: str = "SN"
    returned_device: str = "Returned Device SN"
    pending_return: str = "OLD Device SN"
    enabled: str = ""
    device_allocation: str = "Device(s) Allocation"
    new_asset_status: str = "New Asset Status"


@dataclass(frozen=True)
class SheetRow:
    row_number: int
    deployment_date: date
    username: str | None
    deployment_serial: str | None
    returned_device_serial: str | None
    pending_return_serial: str | None
    marked_red: bool
    enabled: bool
    date_group: int = 1
    returned_device_column_present: bool = True
    device_allocation: str | None = None
    new_asset_status: str | None = None


@dataclass(frozen=True)
class Action:
    group: str
    row_number: int
    username: str
    serial: str
    status: str
    kind: str = "user"
    device_allocation: str | None = None
    new_asset_status: str | None = None
    has_returned_device_serial: bool = False
    has_pending_return_serial: bool = False


def normalized_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def columns_from_mapping(raw: dict[str, Any] | None = None) -> ImportColumns:
    if raw is None:
        return ImportColumns()
    raw = raw or {}
    return ImportColumns(
        username=clean_text(raw.get("username")) or "Username",
        deployment_serial=clean_text(raw.get("deployment_serial")) or "SN",
        returned_device=clean_text(raw.get("returned_device")) or "",
        pending_return=clean_text(raw.get("pending_return")) or "OLD Device SN",
        enabled=clean_text(raw.get("enabled")) or "",
        device_allocation=(
            clean_text(raw.get("device_allocation"))
            if "device_allocation" in raw
            else "Device(s) Allocation"
        ) or "",
        new_asset_status=(
            clean_text(raw.get("new_asset_status"))
            if "new_asset_status" in raw
            else "New Asset Status"
        ) or "",
    )


def find_column_indexes(sheet: Any, columns: ImportColumns) -> tuple[int, dict[str, int], int]:
    """Find a header row by title, allowing the workbook's columns to move."""
    desired = {
        "username": columns.username,
        "deployment_serial": columns.deployment_serial,
        "returned_device": columns.returned_device,
        "pending_return": columns.pending_return,
        "enabled": columns.enabled,
        "device_allocation": columns.device_allocation,
        "new_asset_status": columns.new_asset_status,
    }
    targets = {key: normalized_header(value) for key, value in desired.items()}
    date_titles = {"date", "deployment date", "booking date"}
    max_column = int(sheet.max_column or 1)
    for row in sheet.iter_rows(min_row=1, max_row=min(25, int(sheet.max_row or 25)), max_col=max_column):
        found = {normalized_header(cell.value): cell.column for cell in row if normalized_header(cell.value)}
        indexes = {key: found.get(title) for key, title in targets.items()}
        if not indexes["username"] or not indexes["deployment_serial"] or not indexes["pending_return"]:
            continue
        if not indexes["enabled"] and not columns.enabled:
            # The live ALM sheet normally calls this column "Attend". Use it
            # automatically when no custom TRUE/FALSE heading was configured;
            # otherwise non-attendees look like real users to the warning and
            # import preview.
            for title in (
                "attend",
                "attended",
                "attendance",
                "eligible",
                "enabled",
            ):
                if title in found:
                    indexes["enabled"] = found[title]
                    break
        date_index = next((found[title] for title in date_titles if title in found), None)
        if not date_index:
            continue
        # The returned-device column is intentionally optional until a team adds it.
        return row[0].row, {key: value or 0 for key, value in indexes.items()}, date_index
    required = (
        columns.username,
        columns.deployment_serial,
        columns.pending_return,
        "Date",
    )
    missing = ", ".join(f"{name!r}" for name in required)
    raise eudm.EUDMError(
        "Could not find the required spreadsheet headers. Check ALM Workbook settings: " + missing
    )


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.startswith("#"):
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    return text


def enabled_column_allows(value: Any) -> bool:
    """Treat explicit non-attendance markers as excluded rows."""
    if value is False:
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value != 0
    return str(value).strip().casefold() not in {
        "false",
        "no",
        "n",
        "0",
        "absent",
        "not attended",
        "not attending",
        "no show",
        "no-show",
    }


def looks_like_serial(value: str | None) -> bool:
    """Reject blanks and obvious sheet markers such as 1-5 without overfitting vendors."""
    return is_serial(value)


def looks_like_username(value: str | None) -> bool:
    """Accept a login ID from column D, never a display name or email address."""
    return is_login_id(value)


def cell_is_red(cell: Any) -> bool:
    """Return whether a row cell uses red text as its spreadsheet marker.

    The live Inventory Tracking workbook marks excluded rows with red font,
    not a background fill. Excel may expose that font as ``FF0000``, an ARGB
    value such as ``FFFF0000``/``FFC00000``, or indexed colour 10, so accept
    the common encodings while deliberately ignoring fills.
    """
    color = getattr(getattr(cell, "font", None), "color", None)
    if color is None:
        return False
    color_type = getattr(color, "type", None)
    if color_type == "indexed":
        return getattr(color, "indexed", None) == 10
    if color_type != "rgb":
        return False
    raw = str(getattr(color, "rgb", "") or "").strip().lstrip("#").upper()
    if len(raw) == 8:
        raw = raw[2:]
    if len(raw) != 6:
        return False
    try:
        red, green, blue = (int(raw[index : index + 2], 16) for index in (0, 2, 4))
    except ValueError:
        return False
    return red >= 180 and green <= 100 and blue <= 100


def _color_key(color: Any) -> tuple[Any, ...]:
    """Return a safe, comparable representation of an openpyxl colour."""
    if color is None:
        return (None,)
    values: list[Any] = []
    for name in ("type", "rgb", "indexed", "theme", "tint", "auto"):
        try:
            value = getattr(color, name, None)
        except Exception:
            value = None
        values.append(str(value) if value is not None else None)
    return tuple(values)


def background_fill_key(cell: Any) -> tuple[Any, ...]:
    """Return the background-fill identity used to find date sections.

    The importer intentionally keeps this as workbook metadata rather than
    exposing the actual fill to the web UI. A missing fill is one stable
    section; solid and patterned fills are compared by their type and colours.
    """
    fill = getattr(cell, "fill", None)
    fill_type = str(getattr(fill, "fill_type", None) or "").casefold()
    if not fill_type or fill_type == "none":
        return ("none",)
    return (
        fill_type,
        _color_key(getattr(fill, "fgColor", None)),
        _color_key(getattr(fill, "bgColor", None)),
    )


def normalize_date(value: Any, epoch: datetime) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                pass
    return None


def format_date_label(value: date) -> str:
    """Format an ALM date without platform-specific ``strftime`` directives."""
    return f"{value.strftime('%A')} {value.day} {value.strftime('%B %Y')}"


def username_for(row: tuple[Any, ...], index: int = 4) -> str | None:
    """Read the configured username column only; email is never a fallback."""
    return clean_text(row[index - 1].value) if index else None


def select_workbook_sheet(workbook: Any) -> Any:
    if "Bookings 2026" in workbook.sheetnames:
        return workbook["Bookings 2026"]
    selected = choose_number(
        "The workbook has no sheet named 'Bookings 2026'. Select a sheet",
        list(workbook.sheetnames),
    )
    return workbook[workbook.sheetnames[selected]]


def load_sheet(path: Path, columns: ImportColumns | None = None) -> tuple[str, list[SheetRow]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise eudm.EUDMError(
            "Spreadsheet support is not installed. Run: "
            "python3 -m pip install -r requirements-sheet.txt"
        ) from exc
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except FileNotFoundError as exc:
        raise eudm.EUDMError(f"Spreadsheet not found: {path}") from exc
    except Exception as exc:
        raise eudm.EUDMError(
            f"Could not read {path.name}. Use an unencrypted .xlsx or .xlsm workbook."
        ) from exc
    try:
        sheet = select_workbook_sheet(workbook)
        header_row, indexes, date_index = find_column_indexes(sheet, columns or ImportColumns())
        # Once the headers are mapped, do not stream unrelated trailing
        # columns. Large ALM workbooks often have formatting far beyond the
        # actual table, which otherwise makes every row unnecessarily wide.
        max_column = max(date_index, *indexes.values())
        rows: list[SheetRow] = []
        current_date: date | None = None
        current_fill: tuple[Any, ...] | None = None
        date_group = 0
        for row_offset, values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, min_col=1, max_col=max_column),
            start=1,
        ):
            row_number = getattr(values[0], "row", header_row + row_offset)
            date_cell = values[date_index - 1]
            explicit_date = normalize_date(date_cell.value, workbook.epoch)
            if explicit_date is not None:
                fill_key = background_fill_key(date_cell)
                if explicit_date != current_date:
                    date_group = 1
                elif fill_key != current_fill:
                    date_group += 1
                current_date = explicit_date
                current_fill = fill_key
                deployment_date = explicit_date
            elif current_date is not None and any(
                clean_text(cell.value)
                for index, cell in enumerate(values)
                if index != date_index - 1
            ):
                # Excel often stores a date once for a vertically merged or
                # continued block. Carry the date and Col A fill forward for
                # populated rows beneath it.
                deployment_date = current_date
            else:
                continue
            rows.append(
                SheetRow(
                    row_number=row_number,
                    deployment_date=deployment_date,
                    username=username_for(values, indexes["username"]),
                    deployment_serial=clean_text(values[indexes["deployment_serial"] - 1].value) if indexes["deployment_serial"] else None,
                    returned_device_serial=clean_text(values[indexes["returned_device"] - 1].value) if indexes["returned_device"] else None,
                    pending_return_serial=clean_text(values[indexes["pending_return"] - 1].value) if indexes["pending_return"] else None,
                    marked_red=any(cell_is_red(cell) for cell in values),
                    enabled=enabled_column_allows(
                        values[indexes["enabled"] - 1].value
                    ) if indexes["enabled"] else True,
                    date_group=date_group,
                    returned_device_column_present=bool(indexes["returned_device"]),
                    device_allocation=clean_text(values[indexes["device_allocation"] - 1].value) if indexes["device_allocation"] else None,
                    new_asset_status=clean_text(values[indexes["new_asset_status"] - 1].value) if indexes["new_asset_status"] else None,
                )
            )
        if not rows:
            raise eudm.EUDMError(
                f"No dated data rows were found in columns A-L of sheet {sheet.title!r}."
            )
        return sheet.title, rows
    finally:
        workbook.close()


def latest_download() -> Path | None:
    downloads = Path.home() / "Downloads"
    candidates = [
        path
        for pattern in (f"{FILE_PREFIX}*.xlsx", f"{FILE_PREFIX}*.xlsm")
        for path in downloads.glob(pattern)
        if not path.name.startswith("~$")
    ]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def choose_file(argument: str | None) -> Path:
    if argument:
        return Path(argument).expanduser().resolve()
    latest = latest_download()
    if latest:
        raw = input(f"Spreadsheet [{latest}]: ").strip()
        return Path(raw).expanduser().resolve() if raw else latest.resolve()
    raw = input("Spreadsheet path: ").strip()
    if not raw:
        raise eudm.EUDMError(
            f"No {FILE_PREFIX!r} workbook was found in Downloads and no file was entered."
        )
    return Path(raw).expanduser().resolve()


def choose_number(label: str, options: list[str]) -> int:
    return console.choose_index(label, options)


def yes_no(label: str, *, default: bool = False) -> bool:
    return console.yes_no(label, default=default)


def parse_number_selection(raw: str, maximum: int) -> set[int]:
    selected: set[int] = set()
    if not raw.strip():
        return selected
    for part in raw.split(","):
        token = part.strip()
        match = re.fullmatch(r"(\d+)(?:-(\d+))?", token)
        if not match:
            raise ValueError(f"{token!r} is not a number or range")
        first = int(match.group(1))
        last = int(match.group(2) or first)
        if first > last or first < 1 or last > maximum:
            raise ValueError(f"{token!r} is outside 1-{maximum}")
        selected.update(range(first - 1, last))
    return selected


def eligible_rows(
    rows: Iterable[SheetRow],
    *,
    selected_date: date | None = None,
    date_group: int | None = None,
) -> list[SheetRow]:
    """Return rows that can become imports for the requested date/section.

    Keeping the date and section filtering here prevents summary counters from
    accidentally looking at the whole workbook. A non-empty display name is
    not enough: the importer eventually needs an exact login ID.
    """
    return [
        row
        for row in rows
        if (selected_date is None or row.deployment_date == selected_date)
        and (date_group is None or row.date_group == date_group)
        and row.enabled
        and looks_like_username(row.username)
    ]


def eligible_counts(
    rows: Iterable[SheetRow],
    *,
    selected_date: date | None = None,
    date_group: int | None = None,
) -> tuple[int, int, int]:
    eligible = eligible_rows(
        rows,
        selected_date=selected_date,
        date_group=date_group,
    )
    return (
        sum(looks_like_serial(row.deployment_serial) for row in eligible),
        sum(looks_like_serial(row.returned_device_serial) for row in eligible),
        sum(looks_like_serial(row.pending_return_serial) for row in eligible),
    )


def attended_rows_missing_return_serials(rows: Iterable[SheetRow]) -> list[SheetRow]:
    """Find TRUE/FALSE-enabled rows missing either return serial."""
    return [
        row
        for row in rows
        if row.enabled
        and looks_like_username(row.username)
        and (
            (
                row.returned_device_column_present
                and not looks_like_serial(row.returned_device_serial)
            )
            or not looks_like_serial(row.pending_return_serial)
        )
    ]


def build_actions(
    rows: list[SheetRow],
    selected_date: date,
    mode: str,
    group_selection: str | int | None = None,
) -> tuple[list[Action], Counter[str]]:
    selected_modes = {part.strip() for part in mode.split(",") if part.strip()}
    if "all" in selected_modes:
        selected_modes.update({"deployments", "returned_devices", "pending_returns"})
    selected_group: int | None = None
    raw_group = str(group_selection or "all").strip().casefold()
    if raw_group not in {"", "all"}:
        try:
            selected_group = int(raw_group)
        except ValueError as exc:
            raise eudm.EUDMError("Choose a valid date section.") from exc
        if selected_group < 1:
            raise eudm.EUDMError("Choose a valid date section.")
        available_groups = {
            row.date_group for row in rows if row.deployment_date == selected_date
        }
        if selected_group not in available_groups:
            raise eudm.EUDMError("Choose one of the workbook's available date sections.")
    actions: list[Action] = []
    ignored: Counter[str] = Counter()
    for row in rows:
        if row.deployment_date != selected_date:
            continue
        if selected_group is not None and row.date_group != selected_group:
            continue
        if not row.enabled:
            ignored["TRUE/FALSE column is false"] += 1
            continue
        if not row.username:
            if any(looks_like_serial(value) for value in (row.deployment_serial, row.returned_device_serial, row.pending_return_serial)):
                ignored["Username is blank but serial data is present"] += 1
            continue
        if not looks_like_username(row.username):
            ignored["Username is not a valid login ID"] += 1
            continue
        if "deployments" in selected_modes:
            if looks_like_serial(row.deployment_serial):
                actions.append(Action(
                    "Deployments",
                    row.row_number,
                    row.username,
                    row.deployment_serial,
                    NEW_STOCK,
                    device_allocation=row.device_allocation,
                    new_asset_status=row.new_asset_status,
                    has_returned_device_serial=looks_like_serial(row.returned_device_serial),
                    has_pending_return_serial=looks_like_serial(row.pending_return_serial),
                ))
            else:
                ignored["Deployment serial is blank or invalid"] += 1
        if "returned_devices" in selected_modes:
            if looks_like_serial(row.returned_device_serial):
                actions.append(Action("Returned devices", row.row_number, row.username, row.returned_device_serial, "Used Stock", "location"))
            else:
                ignored["Returned-device serial is blank or invalid"] += 1
        if "pending_returns" in selected_modes:
            if looks_like_serial(row.pending_return_serial):
                actions.append(Action("Pending returns", row.row_number, row.username, row.pending_return_serial, PENDING_RETURN))
            else:
                ignored["Pending-return serial is blank or invalid"] += 1

    serial_spellings: dict[str, str] = {}
    counts: Counter[str] = Counter()
    occurrences: dict[str, list[Action]] = {}
    for action in actions:
        key = action.serial.casefold()
        counts[key] += 1
        serial_spellings.setdefault(key, action.serial)
        occurrences.setdefault(key, []).append(action)
    duplicate_keys = [key for key, count in counts.items() if count > 1]
    if duplicate_keys:
        duplicate_details = []
        for key in duplicate_keys[:12]:
            actions_for_serial = occurrences[key]
            locations = ", ".join(
                f"{action.group} row {action.row_number} ({action.username})"
                for action in actions_for_serial
            )
            duplicate_details.append(f"{serial_spellings[key]} — {locations}")
        extra = len(duplicate_keys) - len(duplicate_details)
        if extra:
            duplicate_details.append(
                f"and {extra} more duplicate serial{'' if extra == 1 else 's'}"
            )
        raise eudm.EUDMError(
            "Duplicate serial numbers were found in the selected date/section. "
            "Each serial can only be imported once; correct the workbook or exclude "
            "one occurrence before continuing: " + "; ".join(duplicate_details)
        )
    return actions, ignored


def override_new_statuses(actions: list[Action]) -> list[Action]:
    new_indexes = [
        index for index, action in enumerate(actions)
        if action.group == "Deployments"
    ]
    if not new_indexes:
        return actions
    print("\nDeployments (default: Deployed - New Stock)")
    for display_index, action_index in enumerate(new_indexes, 1):
        action = actions[action_index]
        print(f"  {display_index}. {action.serial} → {action.username}")
    while True:
        raw = input(
            "Numbers to change to Deployed - Existing Stock "
            "(comma-separated; ranges such as 2-4; Enter for none): "
        )
        try:
            selected = parse_number_selection(raw, len(new_indexes))
            break
        except ValueError as exc:
            print(f"Please correct the selection: {exc}.")
    updated = list(actions)
    for selected_index in selected:
        action_index = new_indexes[selected_index]
        updated[action_index] = replace(updated[action_index], status=EXISTING_STOCK)
    return updated


def print_preview(
    path: Path,
    sheet_name: str,
    selected_date: date,
    actions: list[Action],
    ignored: Counter[str],
) -> None:
    print("\nPreview")
    print(f"  File: {path}")
    print(f"  Sheet: {sheet_name}")
    print(f"  Date: {format_date_label(selected_date)}")
    for group in CLI_GROUPS:
        grouped = [action for action in actions if action.group == group]
        print(f"\n{group} ({len(grouped)})")
        if not grouped:
            print("  None")
        for index, action in enumerate(grouped, 1):
            print(
                f"  {index}. {action.serial} | "
                f"{action.username} | {action.status}"
            )
    if ignored:
        print("\nIgnored serial numbers")
        for reason, count in sorted(ignored.items()):
            print(f"  {count} × {reason}")


def execute(
    client: Any, actions: list[Action], request_for: str, manual_review_enabled: bool, concurrency: int
) -> list[DeploymentOutcome]:
    deployments = [
        UserDeployment(
            serial=action.serial,
            username=action.username,
            status=action.status,
            group=action.group,
            source=f"row {action.row_number}",
        )
        for action in actions
    ]
    return UserDeploymentRunner(
        client, request_for, manual_review=manual_review_enabled, concurrency=concurrency
    ).run(deployments)


def print_results(outcomes: list[DeploymentOutcome]) -> None:
    print_grouped_results(
        outcomes, CLI_GROUPS, command="eudm-inventory-import"
    )


def main() -> int:
    ensure_runtime(requirement_file="requirements-sheet.txt", import_name="openpyxl")
    try:
        config = AppConfig.load()
    except ValueError as exc:
        raise eudm.EUDMError(f"Could not load shared configuration: {exc}") from exc
    if browser_runtime_required(
        sys.argv[1:],
        default_simulate=config.simulate,
        dry_run_flags=("--dry-run",),
    ):
        ensure_runtime(requirement_file="requirements-browser.txt", import_name="playwright")
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Workbook rules:
  The newest 'Inventory Tracking - Sydney*.xlsx' or .xlsm in Downloads is used
  when FILE is omitted. 'Bookings 2026' is selected automatically; otherwise
  you choose a sheet. Columns are identified by their configured headings.
  Non-attendees and rows without a valid username are excluded.
  The standalone importer submits deployments and pending returns; use the web
  importer when returned devices also need to be moved into location stock.

Modes:
  --dry-run ends after the preview with zero browser or EUDM API activity.
  --simulate continues after confirmation using local SIM-REQ/SIM-ORDER IDs,
  but never opens Chrome, reads cookies, contacts EUDM, or changes real data.
  --manual-review asks for a separate y/n approval after each request is populated.
""",
    )
    parser.add_argument("file", nargs="?", metavar="FILE", help="Optional .xlsx/.xlsm workbook. If omitted, choose the newest matching Downloads file or enter a path.")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview only. Does not open Chrome or make any EUDM API requests.",
    )
    parser.add_argument(
        "--request-for",
        default=config.request_for,
        help="Remedy login ID for every request (default: EUDM_REQUEST_FOR).",
    )
    add_runtime_arguments(parser, config)
    args = parser.parse_args()
    validate_runtime_args(args)
    start_run(args, "eudm-inventory-import")

    path = choose_file(args.file)
    sheet_name, rows = load_sheet(path)
    dates = sorted({row.deployment_date for row in rows}, reverse=True)
    date_options = []
    for candidate in dates:
        deployment_count, _, pending_return_count = eligible_counts(
            rows, selected_date=candidate
        )
        date_options.append(
            f"{format_date_label(candidate)} "
            f"({deployment_count} deployments, {pending_return_count} pending returns)"
        )
    selected_date = dates[choose_number("Deployment date", date_options)]
    selected_counts = eligible_counts(rows, selected_date=selected_date)
    mode_index = choose_number(
        "What should be deployed?",
        [
            f"Deployment serials [{selected_counts[0]}]",
            f"Pending-return serials [{selected_counts[2]}]",
            f"Both [{selected_counts[0] + selected_counts[2]}]",
        ],
    )
    mode = (
        "deployments",
        "pending_returns",
        "deployments,pending_returns",
    )[mode_index]
    actions, ignored = build_actions(rows, selected_date, mode)
    if not actions:
        raise eudm.EUDMError("No deployable rows remain for that date and selection.")
    actions = override_new_statuses(actions)
    print_preview(path, sheet_name, selected_date, actions, ignored)

    if args.dry_run:
        print("\nDry run complete. No browser was opened and no EUDM API requests were made.")
        return 0
    requester = resolve_request_for(args, config)
    if not yes_no(f"Submit all {len(actions)} requests now?"):
        print("Cancelled before authentication. No EUDM API requests were made.")
        return 0

    client = open_client(args)
    outcomes = execute(client, actions, requester, args.manual_review, args.concurrency)
    print_results(outcomes)
    return 1 if any(outcome.error for outcome in outcomes) else 0


def cli() -> None:
    """Run the command with stable, user-facing error handling."""
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        raise SystemExit(130)
    except EOFError:
        print("\nInput ended before the import was complete.", file=sys.stderr)
        raise SystemExit(2)
    except eudm.EUDMError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(2)
    except Exception:
        print(
            "Error: An unexpected problem occurred. Re-run with --verbose and "
            "report the last step shown.",
            file=sys.stderr,
        )
        raise SystemExit(2)


if __name__ == "__main__":
    cli()
